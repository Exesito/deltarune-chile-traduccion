#!/usr/bin/env python3
"""
Nucleo compartido para el pipeline de traduccion de Deltarune (Cap.1, solo texto).

Lo usan:
  - builder.py : Sheet descargado (xlsx/csv) -> dist/lang_cl.json + .sha256
  - patcher.py : baja lang_cl.json del repo, valida y parcha el juego

No ejecuta codigo externo: el artefacto de traduccion es SOLO datos (JSON).
"""

import csv
import hashlib
import json
import re
import shutil
from collections import Counter
from pathlib import Path
from datetime import date
from urllib.parse import urljoin, urlparse
from urllib.request import Request, urlopen

BACKUP_SUFFIX = ".orig.bak"

# --- Seguridad: descarga del artefacto de traduccion (lado PATCHER) -------- #
# Hosts desde los que el parchador acepta bajar artefactos publicados.
ALLOWED_HOSTS = {
    "raw.githubusercontent.com",
    "github.com",
    "objects.githubusercontent.com",
    "gitlab.com",
    "codeberg.org",
}
# Hosts del Sheet (lado BUILDER, solo lo corre el mantenedor).
SHEET_HOSTS = {"docs.google.com"}  # + *.googleusercontent.com via _host_ok
MAX_DOWNLOAD_BYTES = 20 * 1024 * 1024  # 20 MB tope duro
MAX_ENTRIES = 200_000                  # tope de claves razonable para Cap.1

# Codigos de control de GameMaker que deben conservarse identicos entre en y cl.
#   \cY \cW ...  -> color (\c + 1 letra); el texto que sigue NO es parte del codigo
#   \M0 \E1 \R   -> codigo (\ + 1 letra + digitos opcionales)
#   ^6           -> timing/pausa      & -> salto de linea     % -> fin
_CODE_RE = re.compile(r"\\c[A-Za-z]|\\[A-Za-z]\d*|\^\d|[&%]")


class SecurityError(Exception):
    """Falla una comprobacion de seguridad; se aborta sin tocar el juego."""


# --------------------------------------------------------------------------- #
# Carga de traduccion (fuente: Sheet descargado)
# --------------------------------------------------------------------------- #
def _pick(header, *names):
    low = [str(h).strip().lower() if h is not None else "" for h in header]
    for n in names:
        if n in low:
            return low.index(n)
    return None


def load_from_json(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_from_xlsx(path: Path) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit(
            "Falta openpyxl para leer .xlsx. Instala: pip install openpyxl\n"
            "(o descarga el Sheet como .csv y usa ese archivo)"
        )
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    table = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0])
        en = row[1] if len(row) > 1 and row[1] is not None else ""
        cl = row[2] if len(row) > 2 and row[2] is not None else ""
        table[key] = {"en": str(en), "cl": str(cl)}
    return table


def rows_to_table(rows: list) -> dict:
    """Convierte filas CSV (con header id/en/cl) a {id: {'en':.., 'cl':..}}."""
    if not rows:
        return {}
    header = rows[0]
    i_id = _pick(header, "id")
    i_id = 0 if i_id is None else i_id
    i_en = _pick(header, "en", "english", "ingles")
    i_en = 1 if i_en is None else i_en
    i_cl = _pick(header, "cl", "es", "espanol", "español", "chile")
    i_cl = 2 if i_cl is None else i_cl
    table = {}
    for r in rows[1:]:
        if not r or i_id >= len(r) or not str(r[i_id]).strip():
            continue
        key = str(r[i_id])
        en = r[i_en] if i_en < len(r) and r[i_en] is not None else ""
        cl = r[i_cl] if i_cl < len(r) and r[i_cl] is not None else ""
        table[key] = {"en": str(en), "cl": str(cl)}
    return table


def load_from_csv(path: Path) -> dict:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        return rows_to_table(list(csv.reader(f)))


def load_csv_bytes(data: bytes) -> dict:
    text = data.decode("utf-8-sig")
    return rows_to_table(list(csv.reader(text.splitlines())))


def load_translation(path: Path) -> dict:
    """Tabla {id: {'en':.., 'cl':..}} desde xlsx/csv/json descargado del Sheet."""
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return load_from_xlsx(path)
    if ext == ".csv":
        return load_from_csv(path)
    if ext == ".json":
        data = load_from_json(path)
        return {k: {"en": v, "cl": v} for k, v in data.items()}
    raise SystemExit(f"Formato no soportado: {ext} (usa .xlsx, .csv o .json)")


def build_lang(table: dict) -> dict:
    """Dict final {id: texto}: usa cl, cae a en si esta vacio."""
    out = {}
    for key, val in table.items():
        cl = (val.get("cl") or "").strip()
        out[key] = cl if cl else val.get("en", "")
    return out


# --------------------------------------------------------------------------- #
# Validacion de codigos de control
# --------------------------------------------------------------------------- #
def validate_codes(table: dict) -> list:
    """Lista de (id, codigos_en, codigos_cl) donde no coinciden (filas traducidas)."""
    issues = []
    for key, val in table.items():
        cl = (val.get("cl") or "").strip()
        if not cl:
            continue
        en_codes = Counter(_CODE_RE.findall(val.get("en", "")))
        cl_codes = Counter(_CODE_RE.findall(cl))
        if en_codes != cl_codes:
            issues.append((key, dict(en_codes), dict(cl_codes)))
    return issues


# --------------------------------------------------------------------------- #
# Integridad (SHA-256)
# --------------------------------------------------------------------------- #
def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def dumps_lang(lang: dict) -> bytes:
    """Serializacion canonica y estable (para que el hash sea reproducible)."""
    return json.dumps(lang, ensure_ascii=False, indent=2, sort_keys=True).encode("utf-8")


# --------------------------------------------------------------------------- #
# Validacion de esquema/seguridad del artefacto de traduccion
# --------------------------------------------------------------------------- #
def validate_lang_schema(obj, reference_keys: set = None) -> list:
    """
    Verifica que 'obj' sea un lang valido: {str: str}, sin estructuras raras.
    Devuelve lista de advertencias (claves faltantes/extra vs referencia).
    Lanza SecurityError si la forma es invalida (posible archivo malicioso/corrupto).
    """
    if not isinstance(obj, dict):
        raise SecurityError("El artefacto no es un objeto JSON {clave: texto}.")
    if len(obj) > MAX_ENTRIES:
        raise SecurityError(f"Demasiadas entradas ({len(obj)} > {MAX_ENTRIES}).")
    for k, v in obj.items():
        if not isinstance(k, str) or not isinstance(v, str):
            raise SecurityError(
                f"Entrada no textual detectada (clave={k!r}). Se esperan solo strings."
            )
    warnings = []
    if reference_keys is not None:
        keys = set(obj.keys())
        missing = reference_keys - keys
        extra = keys - reference_keys
        if missing:
            warnings.append(f"Faltan {len(missing)} claves respecto al lang_en original.")
        if extra:
            warnings.append(
                f"Hay {len(extra)} claves que NO existen en lang_en original "
                f"(ej: {sorted(extra)[:3]}). Sospechoso."
            )
    return warnings


def load_reference_keys(lang_en_path: Path) -> set:
    try:
        return set(load_from_json(lang_en_path).keys())
    except Exception:
        return None


# --------------------------------------------------------------------------- #
# Descarga segura desde el repositorio
# --------------------------------------------------------------------------- #
def _host_ok(hostname: str, allowed: set) -> bool:
    if hostname in allowed:
        return True
    # subdominios de googleusercontent (redireccion del export CSV del Sheet)
    return bool(hostname) and hostname.endswith(".googleusercontent.com")


def fetch_bytes(url: str, allowed_hosts: set = None, allow_any_host: bool = False) -> bytes:
    """Descarga con HTTPS obligatorio, host en allowlist y tope de tamano."""
    allowed = ALLOWED_HOSTS if allowed_hosts is None else allowed_hosts
    parsed = urlparse(url)
    if parsed.scheme != "https":
        raise SecurityError(f"Solo se permite HTTPS (recibido: {parsed.scheme!r}).")
    if not allow_any_host and not _host_ok(parsed.hostname, allowed):
        raise SecurityError(
            f"Host no permitido: {parsed.hostname!r}. Permitidos: {sorted(allowed)}"
        )
    req = Request(url, headers={"User-Agent": "deltarune-cl-patcher"})
    with urlopen(req, timeout=30) as resp:
        data = resp.read(MAX_DOWNLOAD_BYTES + 1)
    if len(data) > MAX_DOWNLOAD_BYTES:
        raise SecurityError(f"Descarga supera el tope de {MAX_DOWNLOAD_BYTES} bytes.")
    return data


# --------------------------------------------------------------------------- #
# Conector al Google Sheet (lado BUILDER, link publico solo-lectura)
# --------------------------------------------------------------------------- #
def sheet_csv_url(sheet_url_or_id: str, gid: str = "0") -> str:
    """
    Construye la URL de export CSV desde una URL de Sheet o un ID pelado.
    Usa el endpoint gviz (docs.google.com, sin redirect a googleusercontent):
    mas confiable que /export?format=csv, que a veces devuelve 400 en su redirect.
    """
    m = re.search(r"/spreadsheets/d/([A-Za-z0-9_-]+)", sheet_url_or_id)
    sheet_id = m.group(1) if m else sheet_url_or_id.strip()
    g = re.search(r"[#?&]gid=(\d+)", sheet_url_or_id)
    if g:
        gid = g.group(1)
    return (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}"
        f"/gviz/tq?tqx=out:csv&gid={gid}"
    )


def fetch_sheet_table(sheet_url_or_id: str, gid: str = "0") -> dict:
    """Baja el Sheet como CSV (link publico) -> tabla {id: {'en','cl'}}."""
    url = sheet_csv_url(sheet_url_or_id, gid)
    data = fetch_bytes(url, allowed_hosts=SHEET_HOSTS)
    if data.lstrip()[:15].lower().startswith(b"<!doctype html") or b"<html" in data[:200].lower():
        raise SecurityError(
            "El Sheet respondio HTML en vez de CSV. Probablemente NO esta compartido "
            "como 'cualquiera con el link: lector'. Ajusta el permiso o usa OAuth."
        )
    return load_csv_bytes(data)


def fetch_lang_from_repo(
    url: str,
    expected_sha256: str = None,
    reference_keys: set = None,
    allow_any_host: bool = False,
) -> tuple:
    """
    Baja lang_cl.json del repo y lo valida antes de devolverlo.
    Devuelve (lang_dict, sha256, warnings).
    """
    data = fetch_bytes(url, allow_any_host=allow_any_host)
    got = sha256_bytes(data)
    if expected_sha256 and got.lower() != expected_sha256.strip().lower():
        raise SecurityError(
            "SHA-256 no coincide.\n"
            f"  esperado: {expected_sha256}\n  recibido: {got}\n"
            "El archivo pudo ser modificado o corromperse. Se aborta."
        )
    try:
        obj = json.loads(data.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as e:
        raise SecurityError(f"El artefacto no es JSON UTF-8 valido: {e}")
    warnings = validate_lang_schema(obj, reference_keys=reference_keys)
    return obj, got, warnings


# --------------------------------------------------------------------------- #
# Parcheo del juego (con contencion de ruta)
# --------------------------------------------------------------------------- #
def find_lang_file(game_path: Path) -> Path:
    """Ubica lang_en.json (archivo directo o dentro de la carpeta del juego)."""
    if game_path.is_file():
        target = game_path
    elif game_path.is_dir():
        candidates = list(game_path.rglob("lang_en.json"))
        if not candidates:
            raise FileNotFoundError(
                f"No encontre lang_en.json dentro de {game_path}. "
                "Apunta directo al archivo lang_en.json."
            )
        target = next(
            (c for c in candidates if c.parent.name.lower() == "lang"), candidates[0]
        )
    else:
        raise FileNotFoundError(f"No existe la ruta: {game_path}")

    # Contencion: el objetivo real debe quedar dentro de la carpeta elegida
    if game_path.is_dir():
        base = game_path.resolve()
        real = target.resolve()
        if base not in real.parents:
            raise SecurityError(
                f"El lang_en.json resuelto ({real}) queda fuera de la carpeta "
                f"del juego ({base}). Posible symlink. Se aborta."
            )
    return target


def patch_game(game_path: Path, lang_data: dict) -> tuple:
    """Respalda lang_en.json (una sola vez) y lo reemplaza con la traduccion."""
    target = find_lang_file(game_path)
    backup = target.with_suffix(target.suffix + BACKUP_SUFFIX)
    backed_up = False
    if not backup.exists():
        shutil.copy2(target, backup)
        backed_up = True
    with open(target, "w", encoding="utf-8") as f:
        json.dump(lang_data, f, ensure_ascii=False, indent=2)
    return target, backup, backed_up


def restore_game(game_path: Path) -> Path:
    target = find_lang_file(game_path)
    backup = target.with_suffix(target.suffix + BACKUP_SUFFIX)
    if not backup.exists():
        raise FileNotFoundError(
            f"No hay respaldo ({backup.name}) junto a {target}. Nada que restaurar."
        )
    shutil.copy2(backup, target)
    return target


# --------------------------------------------------------------------------- #
# Manifest de capitulo (indice de assets: texto, imagenes, fuentes, data.win)
# --------------------------------------------------------------------------- #
def build_manifest(chapter: int, version: str, assets: list) -> dict:
    """assets = [{'type','src','sha256','target','bytes'}, ...]."""
    return {
        "schema": 1,
        "chapter": chapter,
        "version": version or date.today().isoformat(),
        "generated": date.today().isoformat(),
        "assets": assets,
    }


def github_raw_base(owner: str, repo: str, branch: str, subpath: str) -> str:
    sub = subpath.strip("/")
    return f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{sub}/"


def fetch_manifest(base_url: str, allow_any_host: bool = False) -> dict:
    data = fetch_bytes(urljoin(base_url, "manifest.json"), allow_any_host=allow_any_host)
    try:
        man = json.loads(data.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as e:
        raise SecurityError(f"manifest.json invalido: {e}")
    if not isinstance(man, dict) or "assets" not in man:
        raise SecurityError("manifest.json no tiene la forma esperada.")
    return man


def apply_manifest(base_url, manifest, game_path, reference_keys=None,
                   log=print, allow_any_host=False) -> list:
    """
    Descarga y aplica cada asset del manifest, verificando su sha256.
    Hoy soporta 'text'. Los tipos futuros (image/font/data) se avisan y omiten.
    """
    results = []
    for asset in manifest.get("assets", []):
        typ = asset.get("type")
        src = asset.get("src")
        sha = asset.get("sha256")
        if not src:
            continue
        url = urljoin(base_url, src)
        data = fetch_bytes(url, allow_any_host=allow_any_host)
        got = sha256_bytes(data)
        if sha and got.lower() != sha.strip().lower():
            raise SecurityError(
                f"SHA-256 no coincide para {src}.\n  esperado: {sha}\n  recibido: {got}"
            )
        if typ == "text":
            try:
                obj = json.loads(data.decode("utf-8"))
            except (UnicodeDecodeError, json.JSONDecodeError) as e:
                raise SecurityError(f"Asset de texto {src} no es JSON valido: {e}")
            for w in validate_lang_schema(obj, reference_keys=reference_keys):
                log("  aviso: " + w)
            target, backup, backed = patch_game(game_path, obj)
            log(f"  texto aplicado -> {target}" + ("  (respaldo creado)" if backed else ""))
            results.append(("text", str(target)))
        else:
            log(f"  asset '{typ}' aun no soportado por el patcher; se omite ({src}).")
            results.append((typ, "omitido"))
    return results


def format_validation(issues, limit=40) -> list:
    lines = []
    if not issues:
        lines.append("Validacion OK: codigos de control coinciden en todas las filas traducidas.")
        return lines
    lines.append(f"ADVERTENCIA: {len(issues)} fila(s) con codigos distintos entre EN y CL:")
    for key, en_c, cl_c in issues[:limit]:
        lines.append(f"  - {key}: EN={en_c}  CL={cl_c}")
    if len(issues) > limit:
        lines.append(f"  ... y {len(issues) - limit} mas.")
    return lines
